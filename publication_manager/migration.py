from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
import json
from pathlib import Path
import re
import shutil
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from publication_manager.lossless import create_publication_core_from_payload, ensure_template_registry, record_source_row_and_cells
from publication_manager.models import (
    Publication,
    PublicationBookDetails,
    PublicationConferenceDetails,
    PublicationCore,
    PublicationJournalDetails,
    PublicationSourceCell,
    PublicationSourceRow,
)
from publication_manager.normalization import normalize_doi, parse_date


DATE_RANGE_PATTERN = re.compile(r"\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\s+and\s+\d{1,2}[-/]\d{1,2}[-/]\d{2,4}", re.IGNORECASE)
EXPECTED_NAT_VALUES = {"National", "International"}
EXPECTED_TYPES = {"Journal", "Conference", "Book Chapter"}
EXPECTED_CATEGORIES = {"Scopus", "WoS", "UGC Care", "Peer Reviewed", "Book", "International Conference", "National Conference"}

MIGRATION_STATUS_PATH = "migration_status.json"


@dataclass
class SheetConfig:
    category: str
    publication_type: str
    faculty_idx: int
    publication_name_idx: int
    title_idx: int
    authors_idx: int | None
    venue_idx: int | None
    conference_date_idx: int | None
    nat_idx: int | None
    date_idx: int | None
    paper_url_idx: int | None
    doi_idx: int | None
    quartile_idx: int | None
    issn_idx: int | None
    indexing_source: str
    volume_issue_idx: int | None = None
    official_venue_url_idx: int | None = None
    research_published_idx: int | None = None
    indexing_flag_idx: int | None = None
    indexing_proof_idx: int | None = None
    attachment_idx: int | None = None
    presented_accepted_idx: int | None = None
    certificate_idx: int | None = None
    publisher_idx: int | None = None
    book_ugc_idx: int | None = None
    book_scopus_idx: int | None = None
    book_wos_idx: int | None = None


@dataclass
class MigrationReport:
    source_file: str
    db_backup_file: str
    started_at_utc: str
    ended_at_utc: str
    rows_read: int
    rows_imported: int
    rows_skipped: int
    sheet_summary: dict[str, dict[str, int]]
    skip_reasons: dict[str, int]
    quality_checks: list[dict[str, Any]]


SHEET_CONFIGS: dict[str, SheetConfig] = {
    "Scopus Journal": SheetConfig(
        "Scopus",
        "Journal",
        1,
        3,
        5,
        6,
        None,
        None,
        2,
        8,
        11,
        11,
        4,
        14,
        "Scopus",
        volume_issue_idx=7,
        official_venue_url_idx=9,
        research_published_idx=10,
        indexing_flag_idx=12,
        indexing_proof_idx=13,
        attachment_idx=15,
    ),
    "Scopus Conference": SheetConfig(
        "Scopus",
        "Conference",
        1,
        3,
        6,
        7,
        4,
        5,
        2,
        10,
        13,
        13,
        None,
        16,
        "Scopus",
        volume_issue_idx=9,
        official_venue_url_idx=11,
        research_published_idx=12,
        indexing_flag_idx=14,
        indexing_proof_idx=15,
        attachment_idx=18,
        presented_accepted_idx=8,
        certificate_idx=17,
    ),
    "International Conference": SheetConfig(
        "International Conference",
        "Conference",
        1,
        3,
        5,
        6,
        4,
        None,
        2,
        8,
        11,
        11,
        None,
        14,
        "International Conference",
        official_venue_url_idx=9,
        research_published_idx=10,
        indexing_flag_idx=12,
        indexing_proof_idx=13,
        attachment_idx=16,
        presented_accepted_idx=7,
        certificate_idx=15,
    ),
    "National Conference": SheetConfig(
        "National Conference",
        "Conference",
        1,
        3,
        5,
        6,
        4,
        None,
        2,
        8,
        11,
        11,
        None,
        14,
        "National Conference",
        official_venue_url_idx=9,
        research_published_idx=10,
        indexing_flag_idx=12,
        indexing_proof_idx=13,
        attachment_idx=16,
        presented_accepted_idx=7,
        certificate_idx=15,
    ),
    "WoS Journal": SheetConfig(
        "WoS",
        "Journal",
        1,
        3,
        5,
        6,
        None,
        None,
        2,
        8,
        11,
        11,
        4,
        14,
        "WoS",
        volume_issue_idx=7,
        official_venue_url_idx=9,
        research_published_idx=10,
        indexing_flag_idx=12,
        indexing_proof_idx=13,
        attachment_idx=15,
    ),
    "WoS Conference": SheetConfig(
        "WoS",
        "Conference",
        1,
        3,
        5,
        6,
        4,
        None,
        2,
        8,
        11,
        11,
        None,
        14,
        "WoS",
        official_venue_url_idx=9,
        research_published_idx=10,
        indexing_flag_idx=12,
        indexing_proof_idx=13,
        attachment_idx=16,
        presented_accepted_idx=7,
        certificate_idx=15,
    ),
    "Peer Reviewed Journal": SheetConfig(
        "Peer Reviewed",
        "Journal",
        1,
        3,
        4,
        5,
        None,
        None,
        2,
        7,
        10,
        10,
        None,
        11,
        "Peer Reviewed",
        volume_issue_idx=6,
        official_venue_url_idx=8,
        research_published_idx=9,
        attachment_idx=12,
    ),
    "UGC Care Journal": SheetConfig(
        "UGC Care",
        "Journal",
        1,
        3,
        4,
        5,
        None,
        None,
        2,
        7,
        10,
        10,
        None,
        11,
        "UGC Care",
        volume_issue_idx=6,
        official_venue_url_idx=8,
        research_published_idx=9,
        attachment_idx=12,
    ),
    "Book ChapterBook": SheetConfig(
        "Book",
        "Book Chapter",
        1,
        2,
        3,
        4,
        None,
        None,
        None,
        7,
        8,
        8,
        None,
        6,
        "Book",
        official_venue_url_idx=8,
        attachment_idx=12,
        publisher_idx=5,
        book_ugc_idx=9,
        book_scopus_idx=10,
        book_wos_idx=11,
    ),
}


def _sheet_mapping_targets(sheet_name: str) -> dict[int, str]:
    config = SHEET_CONFIGS[sheet_name]
    mappings: dict[int, str] = {1: "source_serial"}

    def put(idx: int | None, target: str) -> None:
        if idx is not None:
            mappings[idx + 1] = target

    put(config.faculty_idx, "faculty_name")
    put(config.publication_name_idx, "publication_name")
    put(config.title_idx, "title")
    put(config.authors_idx, "authors")
    put(config.venue_idx, "venue")
    put(config.conference_date_idx, "conference_date")
    put(config.nat_idx, "national_international")
    put(config.date_idx, "pub_date")
    put(config.paper_url_idx, "paper_url")
    put(config.doi_idx, "doi_source")
    put(config.quartile_idx, "quartile")
    if config.publication_type == "Book Chapter":
        put(config.issn_idx, "isbn")
    else:
        put(config.issn_idx, "issn_isbn")
    put(config.volume_issue_idx, "volume_issue")
    put(config.official_venue_url_idx, "official_venue_url")
    put(config.research_published_idx, "research_published_flag")
    put(config.indexing_flag_idx, "indexing_flag")
    put(config.indexing_proof_idx, "indexing_proof")
    put(config.attachment_idx, "attachment_ref")
    put(config.presented_accepted_idx, "presented_accepted_flag")
    put(config.certificate_idx, "certificate_ref")
    put(config.publisher_idx, "publisher")
    put(config.book_ugc_idx, "book_indexed_ugc")
    put(config.book_scopus_idx, "book_indexed_scopus")
    put(config.book_wos_idx, "book_indexed_wos")
    return mappings


def backup_database(db_path: str, backup_dir: str = "backups") -> str:
    backup_root = Path(backup_dir)
    backup_root.mkdir(parents=True, exist_ok=True)
    source = Path(db_path).resolve()
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    target = backup_root / f"{source.stem}_{stamp}{source.suffix}"
    shutil.copy2(source, target)
    return str(target.resolve())


def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")
    text = str(value).strip()
    return text or None


def _get_col(row: list[Any], idx: int | None) -> Any:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _infer_doi(raw_doi: str | None, paper_url: str | None) -> str | None:
    normalized = normalize_doi(raw_doi)
    if normalized:
        return normalized
    return normalize_doi(paper_url)


def _find_header_row_index(ws) -> int | None:
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        cell_value = ws.cell(row_idx, 1).value
        if isinstance(cell_value, str) and cell_value.strip().lower().startswith("sr"):
            return row_idx
    return None


def _extract_headers(ws) -> list[str]:
    header_row_index = _find_header_row_index(ws)
    if header_row_index is None:
        return []
    headers: list[str] = []
    for col_idx in range(1, ws.max_column + 1):
        raw = ws.cell(header_row_index, col_idx).value
        text = "" if raw is None else str(raw).strip()
        headers.append(text or f"column_{col_idx}")
    return headers


def _data_start_row(ws) -> int:
    header_row_index = _find_header_row_index(ws)
    if header_row_index is None:
        return 1
    return header_row_index + 1


def _persist_publication_records(
    session: Session,
    payload: dict[str, Any],
    source_sheet: str,
    source_row_number: int,
    source_row_values: list[Any],
    sheet_headers: list[str],
    import_batch_id: str,
) -> None:
    session.add(
        Publication(
            faculty_name=payload["faculty_name"],
            title=payload["title"],
            authors=payload["authors"],
            publication_name=payload["publication_name"] or payload["venue"],
            category=payload["category"],
            publication_type=payload["publication_type"],
            venue=payload["venue"],
            conference_date=payload["conference_date"],
            national_international=payload["national_international"],
            pub_date=payload["pub_date"],
            doi=payload["doi"],
            doi_normalized=payload["doi_normalized"],
            paper_url=payload["paper_url"],
            indexing_source=payload["indexing_source"],
            quartile=payload["quartile"],
            issn_isbn=payload["issn_isbn"],
            approved_submission_id=None,
        )
    )
    core_publication = create_publication_core_from_payload(
        session=session,
        payload=payload,
        approved_submission_id=None,
    )
    record_source_row_and_cells(
        session=session,
        publication_id=core_publication.id,
        source_sheet=source_sheet,
        source_row_number=source_row_number,
        row_values=source_row_values,
        headers=sheet_headers,
        import_batch_id=import_batch_id,
        template_version="v1",
    )


def _extract_row_payload(sheet_name: str, row: list[Any]) -> tuple[dict[str, Any] | None, str | None]:
    config = SHEET_CONFIGS[sheet_name]
    if not any(_as_text(value) for value in row):
        return None, "empty_row"

    faculty_name = _as_text(_get_col(row, config.faculty_idx))
    publication_name = _as_text(_get_col(row, config.publication_name_idx))
    title = _as_text(_get_col(row, config.title_idx))
    if not faculty_name or not title:
        return None, "missing_required"

    if config.publication_type == "Conference" and DATE_RANGE_PATTERN.search(title):
        return None, "title_looks_like_date"

    authors = _as_text(_get_col(row, config.authors_idx))
    venue = _as_text(_get_col(row, config.venue_idx))
    publisher = _as_text(_get_col(row, config.publisher_idx))
    conference_date = _as_text(_get_col(row, config.conference_date_idx))
    national_international = _as_text(_get_col(row, config.nat_idx))
    pub_date = parse_date(_get_col(row, config.date_idx))
    paper_url = _as_text(_get_col(row, config.paper_url_idx))
    raw_doi = _as_text(_get_col(row, config.doi_idx))
    doi = _infer_doi(raw_doi, paper_url)
    quartile = _as_text(_get_col(row, config.quartile_idx))
    issn_isbn = _as_text(_get_col(row, config.issn_idx))
    volume_issue = _as_text(_get_col(row, config.volume_issue_idx))
    official_venue_url = _as_text(_get_col(row, config.official_venue_url_idx))
    research_published_flag = _as_text(_get_col(row, config.research_published_idx))
    indexing_flag = _as_text(_get_col(row, config.indexing_flag_idx))
    indexing_proof = _as_text(_get_col(row, config.indexing_proof_idx))
    attachment_ref = _as_text(_get_col(row, config.attachment_idx))
    presented_accepted_flag = _as_text(_get_col(row, config.presented_accepted_idx))
    certificate_ref = _as_text(_get_col(row, config.certificate_idx))
    book_indexed_ugc = _as_text(_get_col(row, config.book_ugc_idx))
    book_indexed_scopus = _as_text(_get_col(row, config.book_scopus_idx))
    book_indexed_wos = _as_text(_get_col(row, config.book_wos_idx))

    if config.publication_type != "Journal":
        quartile = None
    if config.publication_type == "Book Chapter" and not venue and publisher:
        # Keep publisher visible in legacy table/export path while also preserving it losslessly.
        venue = publisher
    if national_international and national_international not in EXPECTED_NAT_VALUES:
        national_international = None

    payload = {
        "faculty_name": faculty_name,
        "publication_name": publication_name,
        "title": title,
        "authors": authors,
        "category": config.category,
        "publication_type": config.publication_type,
        "venue": venue,
        "conference_date": conference_date,
        "national_international": national_international,
        "pub_date": pub_date,
        "doi": doi,
        "doi_normalized": doi,
        "paper_url": paper_url,
        "indexing_source": config.indexing_source,
        "quartile": quartile,
        "issn_isbn": issn_isbn,
        "volume_issue": volume_issue,
        "official_venue_url": official_venue_url,
        "research_published_flag": research_published_flag,
        "indexing_flag": indexing_flag,
        "indexing_proof": indexing_proof,
        "attachment_ref": attachment_ref,
        "presented_accepted_flag": presented_accepted_flag,
        "certificate_ref": certificate_ref,
        "publisher": publisher,
        "isbn": issn_isbn,
        "book_indexed_ugc": book_indexed_ugc,
        "book_indexed_scopus": book_indexed_scopus,
        "book_indexed_wos": book_indexed_wos,
    }
    return payload, None


def run_post_import_quality_checks(session: Session) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []

    total = session.scalar(select(func.count(Publication.id))) or 0
    checks.append({"name": "total_publications_gt_zero", "pass": total > 0, "value": total, "details": "Publications count after import."})

    for col in ("faculty_name", "title", "category", "publication_type"):
        missing = session.scalar(select(func.count(Publication.id)).where((getattr(Publication, col) == None) | (func.trim(getattr(Publication, col)) == ""))) or 0  # noqa: E711
        checks.append({"name": f"required_{col}_complete", "pass": missing == 0, "value": missing, "details": f"Rows missing {col}."})

    conference_titles = session.execute(
        select(Publication.title).where(Publication.publication_type == "Conference")
    ).scalars()
    conference_bad_titles = sum(1 for title in conference_titles if title and DATE_RANGE_PATTERN.search(title))
    checks.append(
        {
            "name": "conference_title_not_date_range",
            "pass": conference_bad_titles == 0,
            "value": conference_bad_titles,
            "details": "Conference titles should not be date ranges.",
        }
    )

    bad_quartile = session.scalar(
        select(func.count(Publication.id)).where(Publication.publication_type != "Journal").where(Publication.quartile.is_not(None))
    ) or 0
    checks.append(
        {
            "name": "quartile_only_for_journal",
            "pass": bad_quartile == 0,
            "value": bad_quartile,
            "details": "Non-journal records should not have quartile.",
        }
    )

    bad_nat = session.scalar(
        select(func.count(Publication.id))
        .where(Publication.national_international.is_not(None))
        .where(Publication.national_international.not_in(EXPECTED_NAT_VALUES))
    ) or 0
    checks.append(
        {
            "name": "national_international_expected_values",
            "pass": bad_nat == 0,
            "value": bad_nat,
            "details": f"Allowed values: {sorted(EXPECTED_NAT_VALUES)}.",
        }
    )

    doi_mismatch = session.scalar(
        select(func.count(Publication.id))
        .where(Publication.doi.is_not(None))
        .where(func.trim(Publication.doi) != "")
        .where((Publication.doi_normalized.is_(None)) | (func.trim(Publication.doi_normalized) == ""))
    ) or 0
    checks.append(
        {
            "name": "doi_normalized_consistency",
            "pass": doi_mismatch == 0,
            "value": doi_mismatch,
            "details": "Records with DOI must have normalized DOI.",
        }
    )
    return checks


def _save_migration_status(status_path: str, report: MigrationReport) -> None:
    payload = asdict(report)
    Path(status_path).write_text(json.dumps(payload, indent=2), encoding="utf-8")


def load_migration_status(status_path: str = MIGRATION_STATUS_PATH) -> dict[str, Any] | None:
    path = Path(status_path)
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def rebuild_publications_from_excel(
    session: Session,
    db_path: str,
    excel_path: str,
    status_path: str = MIGRATION_STATUS_PATH,
) -> MigrationReport:
    started = datetime.now(timezone.utc)
    backup_file = backup_database(db_path)

    session.execute(delete(PublicationSourceCell))
    session.execute(delete(PublicationSourceRow))
    session.execute(delete(PublicationJournalDetails))
    session.execute(delete(PublicationConferenceDetails))
    session.execute(delete(PublicationBookDetails))
    session.execute(delete(PublicationCore))
    session.execute(delete(Publication))
    workbook = load_workbook(Path(excel_path), data_only=True)
    import_batch_id = f"migration:{started.strftime('%Y%m%dT%H%M%SZ')}"

    rows_read = 0
    rows_imported = 0
    rows_skipped = 0
    sheet_summary: dict[str, dict[str, int]] = {}
    skip_reasons: dict[str, int] = {"unsupported_sheet": 0, "invalid_serial": 0, "missing_required": 0, "title_looks_like_date": 0}

    for sheet_name in workbook.sheetnames:
        if sheet_name not in SHEET_CONFIGS:
            skip_reasons["unsupported_sheet"] += 1
            continue
        ws = workbook[sheet_name]
        sheet_headers = _extract_headers(ws)
        start_row = _data_start_row(ws)
        if sheet_headers:
            ensure_template_registry(
                session,
                sheet_name,
                sheet_headers,
                template_version="v1",
                mapping_targets=_sheet_mapping_targets(sheet_name),
            )
        sheet_summary[sheet_name] = {"read": 0, "imported": 0, "skipped": 0}
        for row_number, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            row_values = list(row)
            rows_read += 1
            sheet_summary[sheet_name]["read"] += 1
            payload, reason = _extract_row_payload(sheet_name, row_values)
            if payload is None:
                rows_skipped += 1
                sheet_summary[sheet_name]["skipped"] += 1
                skip_reasons[reason or "missing_required"] = skip_reasons.get(reason or "missing_required", 0) + 1
                continue
            _persist_publication_records(
                session=session,
                payload=payload,
                source_sheet=sheet_name,
                source_row_number=row_number,
                source_row_values=row_values,
                sheet_headers=sheet_headers,
                import_batch_id=import_batch_id,
            )
            rows_imported += 1
            sheet_summary[sheet_name]["imported"] += 1

    quality_checks = run_post_import_quality_checks(session)
    ended = datetime.now(timezone.utc)
    report = MigrationReport(
        source_file=str(Path(excel_path).resolve()),
        db_backup_file=backup_file,
        started_at_utc=started.isoformat(),
        ended_at_utc=ended.isoformat(),
        rows_read=rows_read,
        rows_imported=rows_imported,
        rows_skipped=rows_skipped,
        sheet_summary=sheet_summary,
        skip_reasons=skip_reasons,
        quality_checks=quality_checks,
    )
    _save_migration_status(status_path, report)
    return report


def migrate_from_excel(session: Session, excel_path: str) -> MigrationReport:
    started = datetime.now(timezone.utc)
    workbook = load_workbook(Path(excel_path), data_only=True)
    import_batch_id = f"migration:{started.strftime('%Y%m%dT%H%M%SZ')}"

    rows_read = 0
    rows_imported = 0
    rows_skipped = 0
    sheet_summary: dict[str, dict[str, int]] = {}
    skip_reasons: dict[str, int] = {"unsupported_sheet": 0, "invalid_serial": 0, "missing_required": 0, "title_looks_like_date": 0}

    for sheet_name in workbook.sheetnames:
        if sheet_name not in SHEET_CONFIGS:
            skip_reasons["unsupported_sheet"] += 1
            continue
        ws = workbook[sheet_name]
        sheet_headers = _extract_headers(ws)
        start_row = _data_start_row(ws)
        if sheet_headers:
            ensure_template_registry(
                session,
                sheet_name,
                sheet_headers,
                template_version="v1",
                mapping_targets=_sheet_mapping_targets(sheet_name),
            )
        sheet_summary[sheet_name] = {"read": 0, "imported": 0, "skipped": 0}
        for row_number, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            row_values = list(row)
            rows_read += 1
            sheet_summary[sheet_name]["read"] += 1
            payload, reason = _extract_row_payload(sheet_name, row_values)
            if payload is None:
                rows_skipped += 1
                sheet_summary[sheet_name]["skipped"] += 1
                skip_reasons[reason or "missing_required"] = skip_reasons.get(reason or "missing_required", 0) + 1
                continue
            _persist_publication_records(
                session=session,
                payload=payload,
                source_sheet=sheet_name,
                source_row_number=row_number,
                source_row_values=row_values,
                sheet_headers=sheet_headers,
                import_batch_id=import_batch_id,
            )
            rows_imported += 1
            sheet_summary[sheet_name]["imported"] += 1

    quality_checks = run_post_import_quality_checks(session)
    ended = datetime.now(timezone.utc)
    return MigrationReport(
        source_file=str(Path(excel_path).resolve()),
        db_backup_file="",
        started_at_utc=started.isoformat(),
        ended_at_utc=ended.isoformat(),
        rows_read=rows_read,
        rows_imported=rows_imported,
        rows_skipped=rows_skipped,
        sheet_summary=sheet_summary,
        skip_reasons=skip_reasons,
        quality_checks=quality_checks,
    )
