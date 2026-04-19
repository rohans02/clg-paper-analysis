from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from publication_manager.query import PublicationFilters, get_publications_df


def _build_metadata_df(mode: str, actor: str, row_count: int, filters: dict[str, Any] | None) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "timestamp_utc": datetime.now(timezone.utc).isoformat(),
                "mode": mode,
                "actor": actor,
                "row_count": row_count,
                "filters_json": json.dumps(filters or {}, default=str),
            }
        ]
    )


def _build_xlsx_bytes(data_df: pd.DataFrame, metadata_df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        data_df.to_excel(writer, index=False, sheet_name="publications")
        metadata_df.to_excel(writer, index=False, sheet_name="export_metadata")
    return output.getvalue()


def export_full_xlsx(session: Session, actor: str) -> tuple[bytes, dict[str, Any]]:
    df = get_publications_df(session, PublicationFilters())
    meta = {
        "mode": "full",
        "timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "row_count": len(df),
        "actor": actor,
    }
    metadata_df = _build_metadata_df("full", actor, len(df), None)
    return _build_xlsx_bytes(df, metadata_df), meta


def export_filtered_xlsx(
    session: Session,
    actor: str,
    filters: PublicationFilters,
) -> tuple[bytes, dict[str, Any]]:
    df = get_publications_df(session, filters)
    filter_dict = {
        "faculty_name": filters.faculty_name,
        "category": filters.category,
        "publication_type": filters.publication_type,
        "indexing_source": filters.indexing_source,
        "national_international": filters.national_international,
        "quartile": filters.quartile,
        "keyword": filters.keyword,
        "date_from": filters.date_from,
        "date_to": filters.date_to,
    }
    meta = {
        "mode": "filtered",
        "timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "row_count": len(df),
        "actor": actor,
        "filters": filter_dict,
    }
    metadata_df = _build_metadata_df("filtered", actor, len(df), filter_dict)
    return _build_xlsx_bytes(df, metadata_df), meta


OFFICIAL_SHEET_MAP: dict[tuple[str, str], str] = {
    ("Scopus", "Journal"): "Scopus Journal",
    ("Scopus", "Conference"): "Scopus Conference",
    ("International Conference", "Conference"): "International Conference",
    ("National Conference", "Conference"): "National Conference",
    ("WoS", "Journal"): "WoS Journal",
    ("WoS", "Conference"): "WoS Conference",
    ("Peer Reviewed", "Journal"): "Peer Reviewed Journal",
    ("UGC Care", "Journal"): "UGC Care Journal",
    ("Book", "Book Chapter"): "Book ChapterBook",
}


def _find_data_start_row(ws) -> int:
    header_row = None
    first_numeric = None
    for r in range(1, ws.max_row + 1):
        value = ws.cell(r, 1).value
        if isinstance(value, str) and value.strip().lower().startswith("sr."):
            header_row = r
        if isinstance(value, (int, float)):
            first_numeric = r
            break
    if first_numeric:
        return first_numeric
    if header_row:
        return header_row + 1
    return 4


def _clear_sheet_data(ws, start_row: int) -> None:
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None


def _write_publication_row(ws, sheet_name: str, row_idx: int, serial: int, record: dict[str, Any]) -> None:
    ws.cell(row_idx, 1).value = serial
    ws.cell(row_idx, 2).value = record.get("faculty_name")

    if sheet_name == "Book ChapterBook":
        ws.cell(row_idx, 3).value = record.get("publication_name") or record.get("venue")
        ws.cell(row_idx, 4).value = record.get("title")
        ws.cell(row_idx, 5).value = record.get("authors")
        ws.cell(row_idx, 6).value = record.get("publisher") or record.get("venue")
        ws.cell(row_idx, 7).value = record.get("issn_isbn")
        ws.cell(row_idx, 8).value = record.get("pub_date")
        ws.cell(row_idx, 9).value = record.get("paper_url")
        ws.cell(row_idx, 10).value = record.get("book_indexed_ugc") or "No"
        ws.cell(row_idx, 11).value = record.get("book_indexed_scopus") or ("Yes" if record.get("category") == "Book" else "No")
        ws.cell(row_idx, 12).value = record.get("book_indexed_wos") or "No"
        ws.cell(row_idx, 13).value = record.get("attachment_ref")
        return

    if sheet_name == "Scopus Conference":
        ws.cell(row_idx, 3).value = record.get("national_international")
        ws.cell(row_idx, 4).value = record.get("publication_name") or record.get("venue")
        ws.cell(row_idx, 5).value = record.get("venue")
        ws.cell(row_idx, 6).value = record.get("conference_date")
        ws.cell(row_idx, 7).value = record.get("title")
        ws.cell(row_idx, 8).value = record.get("authors")
        ws.cell(row_idx, 9).value = record.get("presented_accepted_flag")
        ws.cell(row_idx, 10).value = record.get("volume_issue")
        ws.cell(row_idx, 11).value = record.get("pub_date")
        ws.cell(row_idx, 12).value = record.get("official_venue_url")
        ws.cell(row_idx, 13).value = record.get("research_published_flag")
        ws.cell(row_idx, 14).value = record.get("paper_url")
        ws.cell(row_idx, 15).value = record.get("indexing_flag")
        ws.cell(row_idx, 16).value = record.get("indexing_proof")
        ws.cell(row_idx, 17).value = record.get("issn_isbn")
        ws.cell(row_idx, 18).value = record.get("certificate_ref")
        ws.cell(row_idx, 19).value = record.get("attachment_ref")
        return

    if sheet_name in ("International Conference", "National Conference", "WoS Conference"):
        ws.cell(row_idx, 3).value = record.get("national_international")
        ws.cell(row_idx, 4).value = record.get("publication_name") or record.get("venue")
        ws.cell(row_idx, 5).value = record.get("venue")
        ws.cell(row_idx, 6).value = record.get("title")
        ws.cell(row_idx, 7).value = record.get("authors")
        ws.cell(row_idx, 8).value = record.get("presented_accepted_flag")
        ws.cell(row_idx, 9).value = record.get("pub_date")
        ws.cell(row_idx, 10).value = record.get("official_venue_url")
        ws.cell(row_idx, 11).value = record.get("research_published_flag")
        ws.cell(row_idx, 12).value = record.get("paper_url")
        ws.cell(row_idx, 13).value = record.get("indexing_flag")
        ws.cell(row_idx, 14).value = record.get("indexing_proof")
        ws.cell(row_idx, 15).value = record.get("issn_isbn")
        ws.cell(row_idx, 16).value = record.get("certificate_ref")
        ws.cell(row_idx, 17).value = record.get("attachment_ref")
        return

    if sheet_name in ("Scopus Journal", "WoS Journal"):
        ws.cell(row_idx, 3).value = record.get("national_international")
        ws.cell(row_idx, 4).value = record.get("publication_name") or record.get("venue")
        ws.cell(row_idx, 5).value = record.get("quartile")
        ws.cell(row_idx, 6).value = record.get("title")
        ws.cell(row_idx, 7).value = record.get("authors")
        ws.cell(row_idx, 8).value = record.get("volume_issue")
        ws.cell(row_idx, 9).value = record.get("pub_date")
        ws.cell(row_idx, 10).value = record.get("official_venue_url")
        ws.cell(row_idx, 11).value = record.get("research_published_flag")
        ws.cell(row_idx, 12).value = record.get("paper_url")
        ws.cell(row_idx, 13).value = record.get("indexing_flag")
        ws.cell(row_idx, 14).value = record.get("indexing_proof")
        ws.cell(row_idx, 15).value = record.get("issn_isbn")
        ws.cell(row_idx, 16).value = record.get("attachment_ref")
        return

    if sheet_name in ("Peer Reviewed Journal", "UGC Care Journal"):
        ws.cell(row_idx, 3).value = record.get("national_international")
        ws.cell(row_idx, 4).value = record.get("publication_name") or record.get("venue")
        ws.cell(row_idx, 5).value = record.get("title")
        ws.cell(row_idx, 6).value = record.get("authors")
        ws.cell(row_idx, 7).value = record.get("volume_issue")
        ws.cell(row_idx, 8).value = record.get("pub_date")
        ws.cell(row_idx, 9).value = record.get("official_venue_url")
        ws.cell(row_idx, 10).value = record.get("research_published_flag")
        ws.cell(row_idx, 11).value = record.get("paper_url")
        ws.cell(row_idx, 12).value = record.get("issn_isbn")
        ws.cell(row_idx, 13).value = record.get("attachment_ref")


def _populate_analysis_sheet(wb, publications_df: pd.DataFrame) -> None:
    if "Analysis" not in wb.sheetnames:
        return
    ws = wb["Analysis"]

    # Parse existing faculty rows from template.
    faculty_rows: dict[str, int] = {}
    row = 4
    while row <= ws.max_row:
        faculty = ws.cell(row, 2).value
        sr = ws.cell(row, 1).value
        if not faculty and not sr and row > 40:
            break
        if faculty:
            faculty_rows[str(faculty).strip()] = row
        row += 1

    # Include new faculty names not present in template rows.
    all_faculty = sorted(set(publications_df["faculty_name"].dropna().astype(str).str.strip().tolist()))
    next_row = max(faculty_rows.values(), default=3) + 1
    next_sr = max([ws.cell(r, 1).value for r in faculty_rows.values() if isinstance(ws.cell(r, 1).value, (int, float))], default=0) + 1
    for name in all_faculty:
        if name not in faculty_rows:
            ws.cell(next_row, 1).value = next_sr
            ws.cell(next_row, 2).value = name
            faculty_rows[name] = next_row
            next_row += 1
            next_sr += 1

    def count_for(name: str, category: str | None = None, ptype: str | None = None) -> int:
        subset = publications_df[publications_df["faculty_name"] == name]
        if category:
            subset = subset[subset["category"] == category]
        if ptype:
            subset = subset[subset["publication_type"] == ptype]
        return int(len(subset))

    for faculty_name, r in faculty_rows.items():
        ws.cell(r, 3).value = count_for(faculty_name, "Scopus", "Conference")
        ws.cell(r, 4).value = count_for(faculty_name, "Scopus", "Journal")
        ws.cell(r, 5).value = count_for(faculty_name, "WoS", "Journal")
        ws.cell(r, 6).value = count_for(faculty_name, "WoS", "Conference")
        ws.cell(r, 7).value = count_for(faculty_name, "National Conference", "Conference")
        ws.cell(r, 8).value = count_for(faculty_name, "UGC Care", "Journal")
        ws.cell(r, 9).value = count_for(faculty_name, "Peer Reviewed", "Journal")
        ws.cell(r, 10).value = count_for(faculty_name, "Book", "Book Chapter")
        ws.cell(r, 11).value = f"=SUM(C{r}:J{r})"


def export_official_format_xlsx(
    session: Session,
    actor: str,
    template_path: str,
    filters: PublicationFilters | None = None,
) -> tuple[bytes, dict[str, Any]]:
    template_file = Path(template_path).resolve()
    if not template_file.exists():
        raise FileNotFoundError(f"Template file not found: {template_file}")

    publications_df = get_publications_df(session, filters or PublicationFilters())
    wb = load_workbook(template_file)
    start_rows: dict[str, int] = {}

    for (_, _), sheet_name in OFFICIAL_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        start_row = _find_data_start_row(ws)
        start_rows[sheet_name] = start_row
        _clear_sheet_data(ws, start_row)

    for (category, ptype), sheet_name in OFFICIAL_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        subset = publications_df[
            (publications_df["category"] == category)
            & (publications_df["publication_type"] == ptype)
        ].copy()
        if subset.empty:
            continue
        subset = subset.sort_values(["faculty_name", "pub_date", "title"], na_position="last")
        ws = wb[sheet_name]
        start_row = start_rows.get(sheet_name, _find_data_start_row(ws))
        row_idx = start_row
        serial = 1
        for _, row in subset.iterrows():
            _write_publication_row(ws, sheet_name, row_idx, serial, row.to_dict())
            row_idx += 1
            serial += 1

    _populate_analysis_sheet(wb, publications_df)

    output = BytesIO()
    wb.save(output)
    metadata = {
        "mode": "official_format_filtered" if filters else "official_format_full",
        "timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "row_count": len(publications_df),
        "actor": actor,
        "template_path": str(template_file),
        "filters": {
            "faculty_name": filters.faculty_name if filters else None,
            "category": filters.category if filters else None,
            "publication_type": filters.publication_type if filters else None,
            "indexing_source": filters.indexing_source if filters else None,
            "national_international": filters.national_international if filters else None,
            "quartile": filters.quartile if filters else None,
            "keyword": filters.keyword if filters else None,
            "date_from": filters.date_from if filters else None,
            "date_to": filters.date_to if filters else None,
        },
    }
    return output.getvalue(), metadata
