from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from publication_manager.migration import migrate_from_excel
from publication_manager.models import (
    PublicationBookDetails,
    Publication,
    PublicationConferenceDetails,
    PublicationCore,
    PublicationJournalDetails,
    PublicationSourceCell,
    PublicationSourceRow,
    TemplateSchemaRegistry,
)


def test_migration_imports_valid_rows_and_skips_invalid(tmp_path: Path, session):
    file_path = tmp_path / "sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Scopus Journal"
    ws.append(["header"])
    ws.append(["Sr. No.", "Name", "Nat/Int", "Journal", "Q", "Title", "Authors", "V", "Date", "JURL", "Pub", "PaperURL", "Index", "Proof", "ISSN", "Attachment"])
    ws.append([1, "Dr. A", "International", "Journal X", "Q1", "Title X", "A1", "Vol1", "2025-01-01", "https://journal.example", "Yes", "https://doi.org/10.1000/x", "Yes", "https://indexing.example", "1234", "https://attach.example/j"])
    ws.append([2, "", "International", "Journal Y", "Q2", "", "A2", "Vol2", "2025-01-01", "", "", "", "", "", "5678"])

    ws2 = wb.create_sheet("Scopus Conference")
    ws2.append(["header"])
    ws2.append(
        [
            "Sr. No.",
            "Name",
            "Nat/Int",
            "Conference",
            "Venue",
            "Conference Date",
            "Title",
            "Authors",
            "Presented",
            "Vol",
            "Pub Date",
            "Conf URL",
            "Published",
            "Paper URL",
            "Indexing",
            "Proof",
            "ISSN",
            "Certificate",
            "Attachment",
        ]
    )
    ws2.append(
        [
            1,
            "Dr. B",
            "International",
            "IC Test",
            "Online",
            "12-01-2025 and 13-01-2025",
            "Correct Conference Title",
            "A, B",
            "Yes",
            "-",
            "2025-01-20",
            "https://conference.org",
            "Yes",
            "https://doi.org/10.1000/conf.1",
            "Yes",
            "",
            "1234",
            "https://certificate.example/c1",
            "https://attach.example/c1",
        ]
    )
    wb.save(file_path)

    report = migrate_from_excel(session, str(file_path))
    session.flush()
    assert report.rows_imported == 2
    assert report.rows_skipped >= 1
    assert session.query(Publication).count() == 2
    assert session.query(PublicationCore).count() == 2
    assert session.query(PublicationJournalDetails).count() == 1
    assert session.query(PublicationConferenceDetails).count() == 1
    assert session.query(PublicationSourceRow).count() == 2
    assert session.query(PublicationSourceCell).count() > 2
    assert session.query(TemplateSchemaRegistry).count() > 0
    mapped_faculty = (
        session.query(TemplateSchemaRegistry)
        .filter(TemplateSchemaRegistry.sheet_name == "Scopus Journal")
        .filter(TemplateSchemaRegistry.column_index == 2)
        .first()
    )
    assert mapped_faculty is not None
    assert mapped_faculty.mapping_target == "faculty_name"
    conference = session.query(Publication).filter(Publication.publication_type == "Conference").first()
    assert conference is not None
    assert conference.title == "Correct Conference Title"

    journal_details = session.query(PublicationJournalDetails).first()
    assert journal_details is not None
    assert journal_details.volume_issue == "Vol1"
    assert journal_details.official_venue_url == "https://journal.example"
    assert journal_details.research_published_flag == "Yes"
    assert journal_details.indexing_flag == "Yes"
    assert journal_details.indexing_proof == "https://indexing.example"
    assert journal_details.attachment_ref == "https://attach.example/j"

    conference_details = session.query(PublicationConferenceDetails).first()
    assert conference_details is not None
    assert conference_details.presented_accepted_flag == "Yes"
    assert conference_details.volume_issue == "-"
    assert conference_details.official_venue_url == "https://conference.org"
    assert conference_details.research_published_flag == "Yes"
    assert conference_details.indexing_flag == "Yes"
    assert conference_details.certificate_ref == "https://certificate.example/c1"
    assert conference_details.attachment_ref == "https://attach.example/c1"


def test_book_chapter_publisher_and_isbn_are_preserved(tmp_path: Path, session):
    file_path = tmp_path / "book_sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Book ChapterBook"
    ws.append(["Book / Book Chapter"])
    ws.append(
        [
            "Sr. No.",
            "Name of Faculty",
            "Name of Book/ Book Chapter",
            "Title of Book/ Book Chapter",
            "Author(s)",
            "Publisher",
            "ISBN",
            "Date of Publication",
            "URL",
            "UGC",
            "Scopus",
            "WoS",
            "Attachment",
        ]
    )
    ws.append(
        [
            1,
            "Dr. Book",
            "Book Name",
            "Book Chapter Title",
            "A1, A2",
            "Elsevier",
            "978-0-443-36434-1",
            "2026-01-01",
            "https://example.org/book",
            "No",
            "Yes",
            "No",
            "https://attach.example/b1",
        ]
    )
    wb.save(file_path)

    report = migrate_from_excel(session, str(file_path))
    session.flush()

    assert report.rows_imported == 1
    assert session.query(Publication).count() == 1
    assert session.query(PublicationCore).count() == 1
    assert session.query(PublicationBookDetails).count() == 1

    legacy_book = session.query(Publication).first()
    assert legacy_book is not None
    assert legacy_book.venue == "Elsevier"
    assert legacy_book.issn_isbn == "978-0-443-36434-1"

    details = session.query(PublicationBookDetails).first()
    assert details is not None
    assert details.publisher == "Elsevier"
    assert details.isbn == "978-0-443-36434-1"
    assert details.book_indexed_ugc == "No"
    assert details.book_indexed_scopus == "Yes"
    assert details.book_indexed_wos == "No"
    assert details.attachment_ref == "https://attach.example/b1"
    mapped_publisher = (
        session.query(TemplateSchemaRegistry)
        .filter(TemplateSchemaRegistry.sheet_name == "Book ChapterBook")
        .filter(TemplateSchemaRegistry.column_index == 6)
        .first()
    )
    assert mapped_publisher is not None
    assert mapped_publisher.mapping_target == "publisher"
    mapped_isbn = (
        session.query(TemplateSchemaRegistry)
        .filter(TemplateSchemaRegistry.sheet_name == "Book ChapterBook")
        .filter(TemplateSchemaRegistry.column_index == 7)
        .first()
    )
    assert mapped_isbn is not None
    assert mapped_isbn.mapping_target == "isbn"


def test_partial_row_is_kept_but_name_only_is_skipped(tmp_path: Path, session):
    file_path = tmp_path / "eligibility_sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Scopus Journal"
    ws.append(["header"])
    ws.append(["Sr. No.", "Name", "Nat/Int", "Journal", "Q", "Title", "Authors", "V", "Date", "JURL", "Pub", "PaperURL", "Index", "Proof", "ISSN"])

    # Name only -> should be skipped.
    ws.append([3, "Dr. Only Name", None, None, None, None, None, None, None, None, None, None, None, None, None])

    # Missing serial/venue/references but has faculty + title -> should be kept.
    ws.append([None, "Dr. Partial", "International", "Journal Z", None, "Has Core Fields", "A1", None, "2026-01-01", None, None, None, None, None, "ISSN-1234"])

    wb.save(file_path)

    report = migrate_from_excel(session, str(file_path))
    session.flush()

    assert report.rows_imported == 1
    assert report.rows_skipped >= 1
    imported = session.query(Publication).first()
    assert imported is not None
    assert imported.faculty_name == "Dr. Partial"
    assert imported.title == "Has Core Fields"
