from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest
from sqlalchemy.exc import IntegrityError

from publication_manager.enums import InputMethod, SubmissionStatus
from publication_manager.exporter import export_filtered_xlsx, export_full_xlsx, export_official_format_xlsx
from publication_manager.models import PendingSubmission, Publication, PublicationCore, PublicationJournalDetails, PublicationSourceCell, PublicationSourceRow
from publication_manager.query import PublicationFilters
import publication_manager.workflow as workflow_module
from publication_manager.workflow import approve_submission, create_submission, reject_submission


def _payload(doi: str | None = None):
    return {
        "faculty_name": "Dr. A",
        "title": "Paper A",
        "authors": "X, Y",
        "category": "Scopus",
        "publication_type": "Journal",
        "venue": "Test Journal",
        "pub_date": "2025-01-14",
        "doi": doi,
        "paper_url": "https://example.com/paper",
    }


def test_faculty_submit_then_admin_approve_creates_publication(session):
    submission = create_submission(
        session=session,
        submitted_by="faculty1",
        source_input="https://example.com/paper",
        source_input_method=InputMethod.URL,
        payload=_payload("10.1000/abc"),
        confidence_score=0.9,
    )
    session.flush()
    assert submission.status == SubmissionStatus.SUBMITTED.value
    count_before = session.query(Publication).count()
    assert count_before == 0

    result = approve_submission(
        session=session,
        submission_id=submission.id,
        admin_user="admin1",
        review_note="Looks good",
        edited_payload=None,
    )
    session.flush()
    assert result.publication_id is not None
    assert session.query(Publication).count() == 1
    assert session.query(PublicationCore).count() == 1
    assert session.query(PublicationJournalDetails).count() == 1
    assert session.query(PublicationSourceRow).count() == 1
    assert session.query(PublicationSourceCell).count() > 0
    refreshed = session.get(PendingSubmission, submission.id)
    assert refreshed.status == SubmissionStatus.APPROVED.value


def test_faculty_submit_then_admin_reject_creates_no_publication(session):
    submission = create_submission(
        session=session,
        submitted_by="faculty1",
        source_input="manual",
        source_input_method=InputMethod.MANUAL,
        payload=_payload("10.1000/rej"),
        confidence_score=0.7,
    )
    reject_submission(session, submission.id, "admin1", "Incomplete details")
    session.flush()

    assert session.query(Publication).count() == 0
    refreshed = session.get(PendingSubmission, submission.id)
    assert refreshed.status == SubmissionStatus.REJECTED.value


def test_hard_duplicate_blocks_approval(session):
    first = create_submission(
        session=session,
        submitted_by="faculty1",
        source_input="manual",
        source_input_method=InputMethod.MANUAL,
        payload=_payload("10.1000/dup"),
        confidence_score=0.8,
    )
    approve_submission(session, first.id, "admin1")
    second = create_submission(
        session=session,
        submitted_by="faculty2",
        source_input="manual",
        source_input_method=InputMethod.MANUAL,
        payload=_payload("10.1000/dup"),
        confidence_score=0.8,
    )
    result = approve_submission(session, second.id, "admin1")
    assert result.hard_duplicate is True
    assert result.publication_id is None


def test_soft_duplicate_requires_override(session):
    first = create_submission(
        session=session,
        submitted_by="faculty1",
        source_input="manual",
        source_input_method=InputMethod.MANUAL,
        payload=_payload("10.1000/one"),
        confidence_score=0.8,
    )
    approve_submission(session, first.id, "admin1")
    second = create_submission(
        session=session,
        submitted_by="faculty1",
        source_input="manual",
        source_input_method=InputMethod.MANUAL,
        payload=_payload("10.1000/two"),
        confidence_score=0.8,
    )
    result = approve_submission(session, second.id, "admin1")
    assert result.soft_duplicate is True
    assert result.publication_id is None

    result_override = approve_submission(
        session,
        second.id,
        "admin1",
        override_soft_duplicate=True,
        review_note="Valid second publication",
    )
    assert result_override.publication_id is not None


def test_export_full_and_filtered(session):
    s1 = create_submission(
        session=session,
        submitted_by="faculty1",
        source_input="manual",
        source_input_method=InputMethod.MANUAL,
        payload=_payload("10.1000/e1"),
        confidence_score=0.8,
    )
    s2 = create_submission(
        session=session,
        submitted_by="faculty2",
        source_input="manual",
        source_input_method=InputMethod.MANUAL,
        payload={**_payload("10.1000/e2"), "title": "Paper B"},
        confidence_score=0.8,
    )
    approve_submission(session, s1.id, "admin1").publication_id
    approve_submission(session, s2.id, "admin1").publication_id

    full_xlsx, full_meta = export_full_xlsx(session, actor="admin1")
    filtered_xlsx, filtered_meta = export_filtered_xlsx(
        session,
        actor="admin1",
        filters=PublicationFilters(faculty_name="Dr. A"),
    )
    assert full_meta["row_count"] == 2
    assert filtered_meta["row_count"] == 2
    assert len(full_xlsx) > 100
    assert len(filtered_xlsx) > 100


def test_export_official_format_xlsx(session, tmp_path: Path):
    # Arrange approved data.
    s1 = create_submission(
        session=session,
        submitted_by="faculty1",
        source_input="manual",
        source_input_method=InputMethod.MANUAL,
        payload={
            "faculty_name": "Dr. A",
            "title": "Journal Title A",
            "authors": "A1, A2",
            "category": "Scopus",
            "publication_type": "Journal",
            "venue": "Journal X",
            "pub_date": "2025-01-01",
            "paper_url": "https://example.org/j1",
            "quartile": "Q1",
            "national_international": "International",
            "issn_isbn": "1234",
            "volume_issue": "Vol. 1, Issue 1",
            "official_venue_url": "https://journal.example.org",
            "research_published_flag": "Yes",
            "indexing_flag": "Yes",
            "indexing_proof": "https://indexing.example.org/j1",
            "attachment_ref": "https://attachment.example.org/j1",
        },
        confidence_score=0.8,
    )
    approve_submission(session, s1.id, "admin1")

    s2 = create_submission(
        session=session,
        submitted_by="faculty2",
        source_input="manual",
        source_input_method=InputMethod.MANUAL,
        payload={
            "faculty_name": "Dr. B",
            "title": "Conference Title B",
            "authors": "B1, B2",
            "category": "Scopus",
            "publication_type": "Conference",
            "venue": "Conf Y",
            "conference_date": "12-06-2025 and 13-06-2025",
            "pub_date": "2025-02-02",
            "paper_url": "https://example.org/c1",
            "national_international": "International",
            "issn_isbn": "5678",
            "presented_accepted_flag": "Yes",
            "volume_issue": "Volume 88",
            "official_venue_url": "https://conference.example.org",
            "research_published_flag": "Yes",
            "indexing_flag": "In process",
            "indexing_proof": "https://indexing.example.org/c1",
            "certificate_ref": "https://certificate.example.org/c1",
            "attachment_ref": "https://attachment.example.org/c1",
        },
        confidence_score=0.8,
    )
    approve_submission(session, s2.id, "admin1")

    # Build minimal official template.
    template = Workbook()
    ws_j = template.active
    ws_j.title = "Scopus Journal"
    ws_j["A1"] = "SCOPUS Indexed Journal"
    ws_j["A2"] = "Sr. No."
    ws_j["B2"] = "Name of Faculty"
    ws_j["F2"] = "Title of paper"
    ws_j["A4"] = 1
    ws_j["B4"] = "placeholder"

    ws_c = template.create_sheet("Scopus Conference")
    ws_c["A1"] = "SCOPUS Indexed Conference"
    ws_c["A2"] = "Sr. No."
    ws_c["B2"] = "Name of Faculty"
    ws_c["G2"] = "Title of paper"
    ws_c["A3"] = 1
    ws_c["B3"] = "placeholder"

    ws_a = template.create_sheet("Analysis")
    ws_a["A1"] = "Facultywise Publication Analysis"
    ws_a["A2"] = "Sr. No."
    ws_a["B2"] = "Faculty Name"
    ws_a["A4"] = 1
    ws_a["B4"] = "Dr. A"
    ws_a["A5"] = 2
    ws_a["B5"] = "Dr. B"

    template_path = tmp_path / "official_template.xlsx"
    template.save(template_path)

    payload, meta = export_official_format_xlsx(session, "admin1", str(template_path))
    assert meta["mode"] == "official_format_full"
    assert meta["row_count"] == 2

    out_path = tmp_path / "out.xlsx"
    out_path.write_bytes(payload)
    wb = load_workbook(out_path)
    assert "Scopus Journal" in wb.sheetnames
    assert "Scopus Conference" in wb.sheetnames
    assert "Analysis" in wb.sheetnames
    assert wb["Scopus Journal"]["B4"].value == "Dr. A"
    assert wb["Scopus Journal"]["F4"].value == "Journal Title A"
    assert wb["Scopus Journal"]["H4"].value == "Vol. 1, Issue 1"
    assert wb["Scopus Journal"]["J4"].value == "https://journal.example.org"
    assert wb["Scopus Journal"]["K4"].value == "Yes"
    assert wb["Scopus Journal"]["M4"].value == "Yes"
    assert wb["Scopus Journal"]["N4"].value == "https://indexing.example.org/j1"
    assert wb["Scopus Journal"]["P4"].value == "https://attachment.example.org/j1"
    assert wb["Scopus Conference"]["B3"].value == "Dr. B"
    assert wb["Scopus Conference"]["G3"].value == "Conference Title B"
    assert wb["Scopus Conference"]["I3"].value == "Yes"
    assert wb["Scopus Conference"]["J3"].value == "Volume 88"
    assert wb["Scopus Conference"]["L3"].value == "https://conference.example.org"
    assert wb["Scopus Conference"]["M3"].value == "Yes"
    assert wb["Scopus Conference"]["O3"].value == "In process"
    assert wb["Scopus Conference"]["P3"].value == "https://indexing.example.org/c1"
    assert wb["Scopus Conference"]["R3"].value == "https://certificate.example.org/c1"
    assert wb["Scopus Conference"]["S3"].value == "https://attachment.example.org/c1"

    filtered_payload, filtered_meta = export_official_format_xlsx(
        session,
        "admin1",
        str(template_path),
        filters=PublicationFilters(faculty_name="Dr. A"),
    )
    assert filtered_meta["mode"] == "official_format_filtered"
    assert filtered_meta["row_count"] == 1
    filtered_out = tmp_path / "out_filtered.xlsx"
    filtered_out.write_bytes(filtered_payload)
    wb_filtered = load_workbook(filtered_out)
    assert wb_filtered["Scopus Journal"]["B4"].value == "Dr. A"
    assert wb_filtered["Scopus Journal"]["F4"].value == "Journal Title A"
    assert wb_filtered["Scopus Conference"]["B3"].value is None


def test_create_submission_requires_required_fields_for_submitted(session):
    bad_payload = {
        "faculty_name": "Dr. A",
        "category": "Scopus",
        "publication_type": "Journal",
    }
    with pytest.raises(ValueError):
        create_submission(
            session=session,
            submitted_by="faculty1",
            source_input="manual",
            source_input_method=InputMethod.MANUAL,
            payload=bad_payload,
            confidence_score=0.4,
            as_draft=False,
        )

    draft = create_submission(
        session=session,
        submitted_by="faculty1",
        source_input="manual",
        source_input_method=InputMethod.MANUAL,
        payload=bad_payload,
        confidence_score=0.4,
        as_draft=True,
    )
    assert draft.status == SubmissionStatus.DRAFT.value


def test_create_submission_is_idempotent_for_same_payload(session):
    payload = _payload("10.1000/idempotent")
    first = create_submission(
        session=session,
        submitted_by="faculty1",
        source_input="manual",
        source_input_method=InputMethod.MANUAL,
        payload=payload,
        confidence_score=0.8,
    )
    second = create_submission(
        session=session,
        submitted_by="faculty1",
        source_input="manual",
        source_input_method=InputMethod.MANUAL,
        payload=payload,
        confidence_score=0.8,
    )

    assert first.id == second.id
    assert session.query(PendingSubmission).count() == 1


def test_approve_submission_is_idempotent_when_already_approved(session):
    submission = create_submission(
        session=session,
        submitted_by="faculty1",
        source_input="manual",
        source_input_method=InputMethod.MANUAL,
        payload=_payload("10.1000/already-approved"),
        confidence_score=0.8,
    )
    first = approve_submission(session, submission.id, "admin1")
    second = approve_submission(session, submission.id, "admin1")

    assert first.publication_id is not None
    assert second.publication_id == first.publication_id
    assert any("already approved" in msg.lower() for msg in second.warnings)
    assert session.query(Publication).count() == 1


def test_approve_submission_handles_integrity_conflict_without_partial_insert(session, monkeypatch):
    submission = create_submission(
        session=session,
        submitted_by="faculty1",
        source_input="manual",
        source_input_method=InputMethod.MANUAL,
        payload=_payload("10.1000/conflict"),
        confidence_score=0.8,
    )

    def _raise_integrity(*args, **kwargs):
        raise IntegrityError("insert", {}, Exception("forced conflict"))

    monkeypatch.setattr(workflow_module, "create_publication_core_from_payload", _raise_integrity)
    result = approve_submission(session, submission.id, "admin1")

    assert result.publication_id is None
    assert result.hard_duplicate is True
    assert session.query(Publication).count() == 0
    refreshed = session.get(PendingSubmission, submission.id)
    assert refreshed is not None
    assert refreshed.status == SubmissionStatus.UNDER_REVIEW.value
